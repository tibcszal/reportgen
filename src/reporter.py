from typing import Any, Dict, List
import openpyxl as px
import datetime as dt
from openpyxl.utils import get_column_letter


def generate_excel_report(
    analysis_results: List[Dict[str, Any]], output_path: str
) -> None:
    workbook = px.Workbook()

    if not analysis_results:
        active = workbook.active
        assert active is not None
        sheet = active
        sheet.title = "Report"
        sheet.append(["Metric", "Value"])
        sheet.append(["status", "no results"])
    else:
        active = workbook.active
        if active is not None:
            try:
                workbook.remove(active)
            except (ValueError, KeyError):
                pass

        suites: Dict[str, List[Dict[str, Any]]] = {}
        for result in analysis_results:
            test_name: str = str(result.get("test_name", "unknown.unknown"))
            if "." in test_name:
                suite, test = test_name.split(".", 1)
            else:
                suite, test = "__root__", test_name
            grouped = {**result, "_suite": suite, "_test": test}
            suites.setdefault(suite, []).append(grouped)

        summary = workbook.create_sheet(title="Summary", index=0)
        summary.append(
            [
                "Suite",
                "Tests",
                "Error Rate",
                "Avg Resp (ms)",
                "Verdict",
            ]
        )

        overall_tx = 0
        overall_err = 0
        overall_min = None
        overall_max = None
        weighted_sum_avg = 0.0
        weighted_sum_tx = 0
        overall_verdict_fail = False

        for suite_name, results in suites.items():
            suite_tx = 0
            suite_err = 0
            suite_min = None
            suite_max = None
            suite_weighted_sum_avg = 0.0
            suite_weighted_sum_tx = 0
            suite_any_fail = False

            for r in results:
                tx = int(r.get("overall_transaction_count", 0))
                err = int(r.get("overall_error_count", 0))
                rmin = r.get("overall_minimum_response_time")
                rmax = r.get("overall_maximum_response_time")
                avg = float(r.get("overall_avg_response_time", 0.0))
                verdict = str(r.get("verdict", ""))

                suite_tx += tx
                suite_err += err
                if isinstance(rmin, (int, float)):
                    suite_min = rmin if suite_min is None else min(suite_min, rmin)
                if isinstance(rmax, (int, float)):
                    suite_max = rmax if suite_max is None else max(suite_max, rmax)
                suite_weighted_sum_avg += avg * tx
                suite_weighted_sum_tx += tx
                if verdict.upper() == "FAIL":
                    suite_any_fail = True

            suite_err_rate = (suite_err / suite_tx) if suite_tx > 0 else 0.0
            suite_avg = (
                suite_weighted_sum_avg / suite_weighted_sum_tx
                if suite_weighted_sum_tx > 0
                else 0.0
            )
            suite_verdict = "FAIL" if suite_any_fail else "PASS"

            summary.append(
                [
                    suite_name,
                    len(results),
                    suite_err_rate,
                    round(suite_avg, 2),
                    suite_verdict,
                ]
            )
            _r = summary.max_row
            summary.cell(row=_r, column=3).number_format = "0.00%"

            overall_tx += suite_tx
            overall_err += suite_err
            if suite_min is not None:
                overall_min = (
                    suite_min if overall_min is None else min(overall_min, suite_min)
                )
            if suite_max is not None:
                overall_max = (
                    suite_max if overall_max is None else max(overall_max, suite_max)
                )
            weighted_sum_avg += suite_weighted_sum_avg
            weighted_sum_tx += suite_weighted_sum_tx
            overall_verdict_fail = overall_verdict_fail or suite_any_fail

        summary.append([])
        overall_err_rate = (overall_err / overall_tx) if overall_tx > 0 else 0.0
        overall_avg = (
            (weighted_sum_avg / weighted_sum_tx) if weighted_sum_tx > 0 else 0.0
        )
        summary.append(
            [
                "OVERALL",
                sum(len(v) for v in suites.values()),
                overall_err_rate,
                round(overall_avg, 2),
                "FAIL" if overall_verdict_fail else "PASS",
            ]
        )
        _r = summary.max_row
        summary.cell(row=_r, column=3).number_format = "0.00%"

        for suite_name, results in suites.items():
            safe_sheet_name = suite_name[:31] if suite_name else "suite"
            sheet = workbook.create_sheet(title=safe_sheet_name)

            sheet.append(
                [
                    "Test",
                    "Transactions",
                    "Errors",
                    "Error Rate",
                    "Duration (s)",
                    "Avg Resp (ms)",
                    "Min Resp (ms)",
                    "Max Resp (ms)",
                    "Verdict",
                ]
            )

            total_tx = 0
            total_err = 0
            max_of_max = None
            min_of_min = None
            avg_sum = 0.0
            avg_count = 0

            for r in results:
                tx = int(r.get("overall_transaction_count", 0))
                err = int(r.get("overall_error_count", 0))
                dur = int(r.get("test_duration_in_seconds", 0))
                avg = float(r.get("overall_avg_response_time", 0.0))
                rmin = r.get("overall_minimum_response_time")
                rmax = r.get("overall_maximum_response_time")
                verdict = str(r.get("verdict", ""))

                total_tx += tx
                total_err += err
                if isinstance(rmax, (int, float)):
                    max_of_max = rmax if max_of_max is None else max(max_of_max, rmax)
                if isinstance(rmin, (int, float)):
                    min_of_min = rmin if min_of_min is None else min(min_of_min, rmin)
                if isinstance(avg, (int, float)):
                    avg_sum += avg
                    avg_count += 1

                error_rate = (err / tx) if tx > 0 else 0.0
                sheet.append(
                    [
                        r.get("_test", ""),
                        tx,
                        err,
                        error_rate,
                        dur,
                        round(avg, 2),
                        rmin,
                        rmax,
                        verdict,
                    ]
                )
                _rr = sheet.max_row
                sheet.cell(row=_rr, column=4).number_format = "0.00%"

            sheet.append([])
            suite_error_rate = (total_err / total_tx) if total_tx > 0 else 0.0
            suite_avg = (avg_sum / avg_count) if avg_count > 0 else 0.0
            sheet.append(
                [
                    "Suite totals/summary",
                    total_tx,
                    total_err,
                    suite_error_rate,
                    "",
                    round(suite_avg, 2),
                    min_of_min,
                    max_of_max,
                    "",
                ]
            )
            _rr2 = sheet.max_row
            sheet.cell(row=_rr2, column=4).number_format = "0.00%"

            for col_idx, column_cells in enumerate(sheet.columns, start=1):
                max_len = 0
                for c in column_cells:
                    val = c.value
                    if val is None:
                        disp = ""
                    elif (
                        isinstance(val, (int, float))
                        and isinstance(c.number_format, str)
                        and "%" in c.number_format
                    ):
                        try:
                            disp = f"{float(val) * 100:.2f}%"
                        except (ValueError, TypeError):
                            disp = str(val)
                    else:
                        disp = str(val)
                    if len(disp) > max_len:
                        max_len = len(disp)
                width = min(max(max_len + 2, 8), 60)
                sheet.column_dimensions[get_column_letter(col_idx)].width = width

        summary_sheet = workbook["Summary"]
        for col_idx, column_cells in enumerate(summary_sheet.columns, start=1):
            max_len = 0
            for c in column_cells:
                val = c.value
                if val is None:
                    disp = ""
                elif (
                    isinstance(val, (int, float))
                    and isinstance(c.number_format, str)
                    and "%" in c.number_format
                ):
                    try:
                        disp = f"{float(val) * 100:.2f}%"
                    except (ValueError, TypeError):
                        disp = str(val)
                else:
                    disp = str(val)
                if len(disp) > max_len:
                    max_len = len(disp)
            width = min(max(max_len + 2, 8), 60)
            summary_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    workbook.save(
        f"{output_path}/report_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )
