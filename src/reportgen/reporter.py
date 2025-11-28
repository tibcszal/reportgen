from typing import Any, Dict, List
import openpyxl as px
import datetime as dt
from openpyxl.utils import get_column_letter


def generate_excel_report(
    analysis_results: List[Dict[str, Any]], output_path: str
) -> None:
    workbook, err = validate_results(analysis_results)
    if err:
        workbook.save(
            f"{output_path}/no_results_report_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        return
    
    active = workbook.active
    if active is not None:
        try:
            workbook.remove(active)
        except (ValueError, KeyError):
            pass

    suites: Dict[str, Dict[str, Dict[str, Any]]] = {}
    for result in analysis_results:
        test_name: str = str(result.get("test_name", "unknown.unknown"))
        if "." in test_name:
            suite, test = test_name.split(".", 1)
        else:
            suite, test = "__root__", test_name
        base_test = test[:-10] if test.endswith("_resources") else test
        suites.setdefault(suite, {}).setdefault(
            base_test, {"result": None, "resource": None}
        )
        if is_resource_result(result):
            suites[suite][base_test]["resource"] = {**result, "_suite": suite, "_test": base_test}
        else:
            suites[suite][base_test]["result"] = {**result, "_suite": suite, "_test": base_test}

    summary = workbook.create_sheet(title="Summary", index=0)
    summary.append(
        [
            "Suite",
            "Tests",
            "Error Rate",
            "Avg Resp (ms)",
            "Verdict",
            "Avg CPU (mcores)",
            "Max CPU (mcores)",
            "Avg Mem (MiB)",
            "Max Mem (MiB)",
        ]
    )

    overall_tx = 0
    overall_err = 0
    overall_min = None
    overall_max = None
    weighted_sum_avg = 0.0
    weighted_sum_tx = 0
    overall_verdict_fail = False

    overall_cpu_avgs: List[float] = []
    overall_cpu_maxes: List[float] = []
    overall_mem_avgs: List[float] = []
    overall_mem_maxes: List[float] = []

    for suite_name, tests in suites.items():
        suite_tx = 0
        suite_err = 0
        suite_min = None
        suite_max = None
        suite_weighted_sum_avg = 0.0
        suite_weighted_sum_tx = 0
        suite_any_fail = False
        suite_cpu_avgs: List[float] = []
        suite_cpu_maxes: List[float] = []
        suite_mem_avgs: List[float] = []
        suite_mem_maxes: List[float] = []

        for _, group in tests.items():
            r = group.get("result") or {}
            res = group.get("resource") or {}

            tx = int(r.get("overall_transaction_count", 0))
            err = int(r.get("overall_error_count", 0))
            rmin = r.get("overall_minimum_response_time")
            rmax = r.get("overall_maximum_response_time")
            avg = float(r.get("overall_avg_response_time", 0.0))
            verdict = str(r.get("verdict", ""))

            suite_tx += tx
            suite_err += err
            if isinstance(rmin, (int, float)):
                suite_min = rmin if suite_min is None else min(suite_min, rmin)
            if isinstance(rmax, (int, float)):
                suite_max = rmax if suite_max is None else max(suite_max, rmax)
            suite_weighted_sum_avg += avg * max(tx, 0)
            suite_weighted_sum_tx += tx
            if verdict.upper() == "FAIL":
                suite_any_fail = True

            if res:
                cpu_avg, cpu_max, mem_avg, mem_max = extract_resource_overall(res)
                if cpu_avg is not None:
                    suite_cpu_avgs.append(cpu_avg)
                if cpu_max is not None:
                    suite_cpu_maxes.append(cpu_max)
                if mem_avg is not None:
                    suite_mem_avgs.append(mem_avg)
                if mem_max is not None:
                    suite_mem_maxes.append(mem_max)

        suite_err_rate = (suite_err / suite_tx) if suite_tx > 0 else 0.0
        suite_avg = (
            suite_weighted_sum_avg / suite_weighted_sum_tx
            if suite_weighted_sum_tx > 0
            else 0.0
        )
        suite_verdict = "FAIL" if suite_any_fail else "PASS"
        suite_cpu_avg = sum(suite_cpu_avgs) / len(suite_cpu_avgs) if suite_cpu_avgs else None
        suite_cpu_max = max(suite_cpu_maxes) if suite_cpu_maxes else None
        suite_mem_avg = sum(suite_mem_avgs) / len(suite_mem_avgs) if suite_mem_avgs else None
        suite_mem_max = max(suite_mem_maxes) if suite_mem_maxes else None

        summary.append(
            [
                suite_name,
                len(tests),
                suite_err_rate,
                round(suite_avg, 2),
                suite_verdict,
                suite_cpu_avg,
                suite_cpu_max,
                suite_mem_avg,
                suite_mem_max,
            ]
        )
        _r = summary.max_row
        summary.cell(row=_r, column=3).number_format = "0.00%"

        overall_tx += suite_tx
        overall_err += suite_err
        if suite_min is not None:
            overall_min = (
                suite_min if overall_min is None else min(overall_min, suite_min)
            )
        if suite_max is not None:
            overall_max = (
                suite_max if overall_max is None else max(overall_max, suite_max)
            )
        weighted_sum_avg += suite_weighted_sum_avg
        weighted_sum_tx += suite_weighted_sum_tx
        overall_verdict_fail = overall_verdict_fail or suite_any_fail
        if suite_cpu_avg is not None:
            overall_cpu_avgs.append(suite_cpu_avg)
        if suite_cpu_max is not None:
            overall_cpu_maxes.append(suite_cpu_max)
        if suite_mem_avg is not None:
            overall_mem_avgs.append(suite_mem_avg)
        if suite_mem_max is not None:
            overall_mem_maxes.append(suite_mem_max)

    summary.append([])
    overall_err_rate = (overall_err / overall_tx) if overall_tx > 0 else 0.0
    overall_avg = (
        (weighted_sum_avg / weighted_sum_tx) if weighted_sum_tx > 0 else 0.0
    )
    summary.append(
        [
            "OVERALL",
            sum(len(v) for v in suites.values()),
            overall_err_rate,
            round(overall_avg, 2),
            "FAIL" if overall_verdict_fail else "PASS",
            sum(overall_cpu_avgs) / len(overall_cpu_avgs) if overall_cpu_avgs else None,
            max(overall_cpu_maxes) if overall_cpu_maxes else None,
            sum(overall_mem_avgs) / len(overall_mem_avgs) if overall_mem_avgs else None,
            max(overall_mem_maxes) if overall_mem_maxes else None,
        ]
    )
    _r = summary.max_row
    summary.cell(row=_r, column=3).number_format = "0.00%"

    for suite_name, tests in suites.items():
        safe_sheet_name = suite_name[:31] if suite_name else "suite"
        sheet = workbook.create_sheet(title=safe_sheet_name)

        sheet.append(
            [
                "Test",
                "Transactions",
                "Errors",
                "Error Rate",
                "Duration (s)",
                "Avg Resp (ms)",
                "Min Resp (ms)",
                "Max Resp (ms)",
                "Verdict",
                "Avg CPU (mcores)",
                "Max CPU (mcores)",
                "Avg Mem (MiB)",
                "Max Mem (MiB)",
            ]
        )

        total_tx = 0
        total_err = 0
        max_of_max = None
        min_of_min = None
        avg_sum = 0.0
        avg_count = 0
        suite_cpu_avgs_local: List[float] = []
        suite_cpu_maxes_local: List[float] = []
        suite_mem_avgs_local: List[float] = []
        suite_mem_maxes_local: List[float] = []

        for _, group in tests.items():
            r = group.get("result") or {}
            res = group.get("resource") or {}

            tx = int(r.get("overall_transaction_count", 0))
            err = int(r.get("overall_error_count", 0))
            dur = int(r.get("test_duration_in_seconds", 0))
            avg = float(r.get("overall_avg_response_time", 0.0))
            rmin = r.get("overall_minimum_response_time")
            rmax = r.get("overall_maximum_response_time")
            verdict = str(r.get("verdict", ""))
            cpu_avg, cpu_max, mem_avg, mem_max = extract_resource_overall(res)

            total_tx += tx
            total_err += err
            if isinstance(rmax, (int, float)):
                max_of_max = rmax if max_of_max is None else max(max_of_max, rmax)
            if isinstance(rmin, (int, float)):
                min_of_min = rmin if min_of_min is None else min(min_of_min, rmin)
            if isinstance(avg, (int, float)):
                avg_sum += avg
                avg_count += 1
            if cpu_avg is not None:
                suite_cpu_avgs_local.append(cpu_avg)
            if cpu_max is not None:
                suite_cpu_maxes_local.append(cpu_max)
            if mem_avg is not None:
                suite_mem_avgs_local.append(mem_avg)
            if mem_max is not None:
                suite_mem_maxes_local.append(mem_max)

            error_rate = (err / tx) if tx > 0 else 0.0
            sheet.append(
                [
                    r.get("_test", "") or res.get("_test", ""),
                    tx,
                    err,
                    error_rate,
                    dur,
                    round(avg, 2),
                    rmin,
                    rmax,
                    verdict,
                    cpu_avg,
                    cpu_max,
                    mem_avg,
                    mem_max,
                ]
            )
            _rr = sheet.max_row
            sheet.cell(row=_rr, column=4).number_format = "0.00%"

        sheet.append([])
        suite_error_rate = (total_err / total_tx) if total_tx > 0 else 0.0
        suite_avg = (avg_sum / avg_count) if avg_count > 0 else 0.0
        suite_cpu_avg_val = (
            sum(suite_cpu_avgs_local) / len(suite_cpu_avgs_local)
            if suite_cpu_avgs_local
            else None
        )
        suite_cpu_max_val = max(suite_cpu_maxes_local) if suite_cpu_maxes_local else None
        suite_mem_avg_val = (
            sum(suite_mem_avgs_local) / len(suite_mem_avgs_local)
            if suite_mem_avgs_local
            else None
        )
        suite_mem_max_val = max(suite_mem_maxes_local) if suite_mem_maxes_local else None
        sheet.append(
            [
                "Suite totals/summary",
                total_tx,
                total_err,
                suite_error_rate,
                "",
                round(suite_avg, 2),
                min_of_min,
                max_of_max,
                "",
                suite_cpu_avg_val,
                suite_cpu_max_val,
                suite_mem_avg_val,
                suite_mem_max_val,
            ]
        )
        _rr2 = sheet.max_row
        sheet.cell(row=_rr2, column=4).number_format = "0.00%"

        for col_idx, column_cells in enumerate(sheet.columns, start=1):
            max_len = 0
            for c in column_cells:
                val = c.value
                if val is None:
                    disp = ""
                elif (
                    isinstance(val, (int, float))
                    and isinstance(c.number_format, str)
                    and "%" in c.number_format
                ):
                    try:
                        disp = f"{float(val) * 100:.2f}%"
                    except (ValueError, TypeError):
                        disp = str(val)
                else:
                    disp = str(val)
                if len(disp) > max_len:
                    max_len = len(disp)
            width = min(max(max_len + 2, 8), 60)
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

    summary_sheet = workbook["Summary"]
    for col_idx, column_cells in enumerate(summary_sheet.columns, start=1):
        max_len = 0
        for c in column_cells:
            val = c.value
            if val is None:
                disp = ""
            elif (
                isinstance(val, (int, float))
                and isinstance(c.number_format, str)
                and "%" in c.number_format
            ):
                try:
                    disp = f"{float(val) * 100:.2f}%"
                except (ValueError, TypeError):
                    disp = str(val)
            else:
                disp = str(val)
            if len(disp) > max_len:
                max_len = len(disp)
        width = min(max(max_len + 2, 8), 60)
        summary_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    workbook.save(
        f"{output_path}/report_{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    
def validate_results(analysis_results: List[Dict[str, Any]]) -> tuple[px.Workbook, str]:
    workbook = px.Workbook()
    if not analysis_results:
        active = workbook.active
        assert active is not None
        sheet = active
        sheet.title = "Report"
        sheet.append(["Metric", "Value"])
        sheet.append(["status", "no results"])
        return workbook, "no results"
    return workbook, ""


def is_resource_result(result: Dict[str, Any]) -> bool:
    return "cpu_avg_per_pod" in result or "overall_avg_cpu_mcores" in result.get("overall", {})


def extract_resource_overall(result: Dict[str, Any]) -> tuple[Any, Any, Any, Any]:
    if not result:
        return None, None, None, None
    overall = result.get("overall", {})
    cpu_avg = overall.get("overall_avg_cpu_mcores")
    cpu_max = overall.get("overall_max_cpu_mcores")
    mem_avg_bytes = overall.get("overall_avg_memory_bytes")
    mem_max_bytes = overall.get("overall_max_memory_bytes")
    mem_avg_mib = bytes_to_mib(mem_avg_bytes) if mem_avg_bytes is not None else None
    mem_max_mib = bytes_to_mib(mem_max_bytes) if mem_max_bytes is not None else None
    return cpu_avg, cpu_max, mem_avg_mib, mem_max_mib


def bytes_to_mib(value: Any) -> float | None:
    if value is None:
        return None
    try:
        return float(value) / (1024**2)
    except (TypeError, ValueError):
        return None
